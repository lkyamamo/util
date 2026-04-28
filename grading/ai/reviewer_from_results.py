#!/usr/bin/env python3
"""
Reviewer UI driven by grader output: separate written and code results CSVs.

Usage:
    python reviewer_from_results.py <directory>

The directory must contain:
  - exactly one file matching *_written.csv
  - exactly one file matching *_code.csv
  - exactly one file matching rubric*.json (or *.json if none start with "rubric")
  - exactly one subdirectory matching *submissions*

Example:
    python reviewer_from_results.py /path/to/assignment7/
"""

import argparse
import os
import re
import csv
import json
import urllib.parse
from pathlib import Path
from http.server import HTTPServer, SimpleHTTPRequestHandler
import webbrowser
import threading
import sys

try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

# Extracts "Sydney Leggio" from "362403-553690 - Sydney Leggio - Apr 6, 2026 1239 PM"
_NAME_FROM_FOLDER = re.compile(r'^\d[\w-]*\s+-\s+(.+?)\s+-\s+\w+\s+\d+,', re.I)

# Match "Problem N" or "for Problem N" etc. in requirement text
_PROBLEM_RE = re.compile(r"\bProblem\s*(\d+)\b", re.I)


def _problem_label(requirement_text):
    """Extract problem label from requirement text, e.g. 'Problem 1' or 'General'."""
    if not requirement_text:
        return "General"
    m = _PROBLEM_RE.search(requirement_text)
    return f"Problem {m.group(1)}" if m else "General"


# ----------------------------
# Data Loader
# ----------------------------

def _parse_results_csv(csv_path, skip_columns=frozenset({"Student", "Needs_Review"})):
    """Load a single results CSV; return (requirement_columns, list of (student, warnings, missing))."""
    path = Path(csv_path)
    if not path.exists():
        return [], []

    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    if not rows:
        return [], []

    columns = [c for c in rows[0].keys() if c not in skip_columns and (rows[0].get(c) or "").strip().upper() not in ("TRUE", "FALSE")]

    out = []
    for row in rows:
        student = (row.get("Student") or "").strip()
        if not student:
            continue
        warnings = []
        missing = []
        for col in columns:
            status = (row.get(col) or "").strip().upper()
            if status in ("UNCERTAIN", "NEEDS_REVIEW"):
                warnings.append(col)
            elif status == "NOT_MET":
                missing.append(col)
        out.append((student, warnings, missing))

    return columns, out


class FeedbackStore:
    """Persist per-student comments and points to a CSV file."""

    FIELDNAMES = ["Student", "Points", "Comments"]

    def __init__(self, csv_path):
        self.csv_path = Path(csv_path)
        self._lock = threading.Lock()
        self._data = {}  # {student: {points: str, comments: str}}
        self._load()

    def _load(self):
        if not self.csv_path.exists():
            return
        with open(self.csv_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                student = (row.get("Student") or "").strip()
                if student:
                    self._data[student] = {
                        "points": (row.get("Points") or "").strip(),
                        "comments": (row.get("Comments") or "").strip(),
                    }

    def _save(self):
        with open(self.csv_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=self.FIELDNAMES)
            writer.writeheader()
            for student in sorted(self._data):
                writer.writerow({
                    "Student": student,
                    "Points": self._data[student].get("points", ""),
                    "Comments": self._data[student].get("comments", ""),
                })

    def get(self, student):
        with self._lock:
            return dict(self._data.get(student, {"points": "", "comments": ""}))

    def set(self, student, points, comments):
        with self._lock:
            self._data[student] = {"points": points, "comments": comments}
            self._save()


class ResultsParser:
    def __init__(self, written_csv, code_csv, rubric_path, submissions_root, feedback_csv,
                 grading_xlsx=None, assignment_num=None):
        self.written_csv = Path(written_csv)
        self.code_csv = Path(code_csv)
        self.rubric_path = Path(rubric_path)
        self.submissions_root = Path(submissions_root)
        self.feedback = FeedbackStore(feedback_csv)
        self.grading_xlsx = Path(grading_xlsx) if grading_xlsx else None
        self.assignment_num = assignment_num  # int, e.g. 9

        self.students = {}
        self.requirements = {}
        self.written_columns = []
        self.code_columns = []

    def parse(self):
        # Optional rubric for requirement descriptions
        if self.rubric_path.exists():
            with open(self.rubric_path, "r") as f:
                rubric = json.load(f)
            if "requirements" in rubric:
                for r in rubric["requirements"]:
                    self.requirements[r["id"]] = r["description"]
            else:
                for key in ("written_requirements", "code_requirements", "integration_requirements"):
                    for desc in rubric.get(key, []):
                        self.requirements[desc] = desc

        written_cols, written_rows = _parse_results_csv(self.written_csv)
        code_cols, code_rows = _parse_results_csv(self.code_csv)

        for c in written_cols + code_cols:
            if c not in self.requirements:
                self.requirements[c] = c

        self.written_columns = written_cols
        self.code_columns = code_cols

        by_student_written = {student: (warnings, missing) for student, warnings, missing in written_rows}
        by_student_code = {student: (warnings, missing) for student, warnings, missing in code_rows}

        all_students = set(by_student_written) | set(by_student_code)
        for student in all_students:
            w_warn, w_miss = by_student_written.get(student, ([], []))
            c_warn, c_miss = by_student_code.get(student, ([], []))
            self.students[student] = {
                "dir_name": student,
                "written": {"warnings": w_warn, "missing": w_miss},
                "code": {"warnings": c_warn, "missing": c_miss},
                "warnings": w_warn + c_warn,
                "missing": w_miss + c_miss,
            }

    def find_pdf_files(self, student_name):
        student_dir = self.submissions_root / student_name
        if not student_dir.exists():
            return []

        pdf_files = []
        for pdf_file in student_dir.rglob("*.pdf"):
            rel_path = pdf_file.relative_to(self.submissions_root)
            pdf_files.append({
                "name": pdf_file.name,
                "path": str(rel_path),
                "full_path": str(pdf_file)
            })

        return sorted(pdf_files, key=lambda x: x["name"])

    def export_to_excel(self):
        """Write saved feedback to an xlsx file. Creates a new file if none exists."""
        if not _OPENPYXL_AVAILABLE:
            return False, "openpyxl is not installed (pip install openpyxl)"

        if not self.feedback._data:
            return False, "No feedback data to export."

        output_path = self.grading_xlsx
        if not output_path:
            output_path = self.feedback.csv_path.with_suffix(".xlsx")

        if output_path.exists() and self.assignment_num:
            return self._merge_into_existing_xlsx(output_path)

        return self._create_new_xlsx(output_path)

    def _create_new_xlsx(self, output_path):
        """Create a standalone xlsx from feedback data."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grades"

        ws.cell(1, 1).value = "Name"
        ws.cell(1, 2).value = "Points"
        ws.cell(1, 3).value = "Comments"

        row = 2
        for folder_name, entry in sorted(self.feedback._data.items()):
            m = _NAME_FROM_FOLDER.match(folder_name)
            student_name = m.group(1).strip() if m else folder_name.strip()

            ws.cell(row, 1).value = student_name
            points = entry.get("points", "").strip()
            if points:
                try:
                    ws.cell(row, 2).value = float(points)
                except ValueError:
                    ws.cell(row, 2).value = points
            ws.cell(row, 3).value = entry.get("comments", "").strip()
            row += 1

        wb.save(str(output_path))
        return True, f"Exported {row - 2} student(s) to {output_path.name}."

    def _merge_into_existing_xlsx(self, output_path):
        """Merge feedback into an existing grading xlsx."""
        point_col_name = f"Sheet {self.assignment_num} Point Grade"
        text_col_name  = f"Sheet {self.assignment_num} Text Grade"

        wb = openpyxl.load_workbook(str(output_path))
        ws = wb.active

        header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_map = {str(v).strip(): i + 1 for i, v in enumerate(header_row) if v is not None}

        if point_col_name not in col_map:
            next_col = ws.max_column + 1
            ws.cell(1, next_col).value = point_col_name
            col_map[point_col_name] = next_col
        if text_col_name not in col_map:
            next_col = ws.max_column + 1
            ws.cell(1, next_col).value = text_col_name
            col_map[text_col_name] = next_col

        point_col = col_map[point_col_name]
        text_col  = col_map[text_col_name]

        name_col = col_map.get("Name", 1)
        name_to_row = {}
        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row, name_col).value
            if cell_val:
                name_to_row[str(cell_val).strip().lower()] = row

        written = 0
        unmatched = []
        for folder_name, entry in self.feedback._data.items():
            m = _NAME_FROM_FOLDER.match(folder_name)
            student_name = m.group(1).strip() if m else folder_name.strip()
            key = student_name.lower()

            if key not in name_to_row:
                key = next((k for k in name_to_row if key in k or k in key), None)

            if not key or key not in name_to_row:
                unmatched.append(folder_name)
                continue

            row = name_to_row[key]
            points = entry.get("points", "").strip()
            comments = entry.get("comments", "").strip()
            if points:
                try:
                    ws.cell(row, point_col).value = float(points)
                except ValueError:
                    ws.cell(row, point_col).value = points
            if comments:
                ws.cell(row, text_col).value = comments
            written += 1

        wb.save(str(output_path))

        msg = f"Exported {written} student(s) to {output_path.name}."
        if unmatched:
            msg += f" Could not match: {', '.join(unmatched)}"
        return True, msg


# ----------------------------
# HTTP Handler
# ----------------------------

class ReviewerHandler(SimpleHTTPRequestHandler):

    def __init__(self, *args, parser=None, **kwargs):
        self.parser = parser
        super().__init__(*args, **kwargs)

    def do_GET(self):

        if self.path == "/" or self.path == "/index.html":
            self._send_html(self.generate_html())

        elif self.path.startswith("/api/students"):
            self._send_json(self.get_students_data())

        elif self.path.startswith("/api/student/"):
            student_name = urllib.parse.unquote(
                self.path.split("/api/student/")[-1]
            )
            self._send_json(self.get_student_data(student_name))

        elif self.path.startswith("/api/feedback/"):
            student_name = urllib.parse.unquote(
                self.path.split("/api/feedback/")[-1]
            )
            self._send_json(self.parser.feedback.get(student_name))

        elif self.path.startswith("/pdf/"):
            self._serve_pdf()

        else:
            super().do_GET()

    def do_POST(self):
        if self.path == "/api/feedback":
            length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(length)
            try:
                payload = json.loads(body.decode("utf-8"))
                student = (payload.get("student") or "").strip()
                points = str(payload.get("points") or "").strip()
                comments = str(payload.get("comments") or "").strip()
                if not student:
                    self._send_json({"ok": False, "error": "missing student"})
                    return
                self.parser.feedback.set(student, points, comments)
                self._send_json({"ok": True})
            except Exception as e:
                self._send_json({"ok": False, "error": str(e)})
        elif self.path == "/api/export":
            try:
                ok, msg = self.parser.export_to_excel()
                self._send_json({"ok": ok, "message": msg})
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)})
        else:
            self.send_response(404)
            self.end_headers()

    def _send_html(self, html):
        self.send_response(200)
        self.send_header("Content-type", "text/html")
        self.end_headers()
        self.wfile.write(html.encode("utf-8"))

    def _send_json(self, data):
        self.send_response(200)
        self.send_header("Content-type", "application/json")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(json.dumps(data).encode("utf-8"))

    def _serve_pdf(self):
        pdf_path_raw = self.path[5:]
        pdf_path = urllib.parse.unquote(pdf_path_raw)
        full_path = self.parser.submissions_root / pdf_path

        if full_path.exists() and full_path.suffix == ".pdf":
            self.send_response(200)
            self.send_header("Content-type", "application/pdf")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            with open(full_path, "rb") as f:
                self.wfile.write(f.read())
        else:
            self.send_response(404)
            self.end_headers()

    def get_students_data(self):
        return {
            "students": [
                {
                    "name": name,
                    "warnings_count": len(data["warnings"]),
                    "missing_count": len(data["missing"]),
                    "written_warnings": len(data["written"]["warnings"]),
                    "written_missing": len(data["written"]["missing"]),
                    "code_warnings": len(data["code"]["warnings"]),
                    "code_missing": len(data["code"]["missing"]),
                }
                for name, data in sorted(self.parser.students.items())
            ]
        }

    def get_student_data(self, student_name):
        if student_name not in self.parser.students:
            return {"error": "Student not found"}

        data = self.parser.students[student_name]
        pdf_files = self.parser.find_pdf_files(student_name)
        req_desc = self.parser.requirements.get

        def details(req_ids):
            return [
                {
                    "req_id": r,
                    "description": req_desc(r, r),
                    "problem": _problem_label(r),
                }
                for r in req_ids
            ]

        return {
            "name": student_name,
            "written": {
                "warnings": details(data["written"]["warnings"]),
                "missing": details(data["written"]["missing"]),
            },
            "code": {
                "warnings": details(data["code"]["warnings"]),
                "missing": details(data["code"]["missing"]),
            },
            "pdf_files": pdf_files,
        }

    def generate_html(self):
        return """
<!DOCTYPE html>
<html>
<head>
<title>AI Grading Reviewer</title>
<style>
* { box-sizing: border-box; }
body { font-family: Arial; display: flex; height: 100vh; margin: 0; overflow: hidden; }

/* ── Left sidebar: student list ── */
.sidebar {
  width: 280px;
  min-width: 220px;
  background: #f5f5f5;
  overflow-y: auto;
  padding: 10px;
  border-right: 1px solid #ddd;
  flex-shrink: 0;
}
.sidebar h2 { margin: 0 0 10px 0; font-size: 1rem; color: #333; }
.student { padding: 8px; cursor: pointer; border-bottom: 1px solid #ddd; }
.student:hover { background: #e8e8e8; }
.student.selected { background: #d0e8ff; }

/* ── Middle: PDF + checklist ── */
.main {
  flex: 1;
  display: flex;
  flex-direction: column;
  min-width: 0;
  overflow: hidden;
}
.viewer { flex: 2; min-height: 0; }
.viewer iframe { width: 100%; height: 100%; border: none; }
.checklist {
  flex: 1;
  padding: 16px;
  overflow-y: auto;
  border-top: 1px solid #ddd;
}
.checklist-section { margin-bottom: 24px; }
.checklist-section h3 { margin: 0 0 10px 0; font-size: 1rem; color: #333; }
.problem-group { margin-bottom: 16px; }
.problem-group h4 { margin: 0 0 8px 0; font-size: 0.9rem; color: #555; font-weight: 600; }
.problem-badge { display: inline-block; background: #e0e0e0; color: #333; padding: 2px 8px; border-radius: 4px; font-size: 0.8rem; margin-right: 8px; }
.missing { border-left: 4px solid #c00; padding: 10px; margin-bottom: 10px; background: #fff5f5; }
.warning { border-left: 4px solid #888; padding: 10px; margin-bottom: 10px; background: #f5f5f5; }
.no-issues { color: #666; font-style: italic; }

/* ── Right panel: feedback ── */
.feedback-panel {
  width: 300px;
  min-width: 240px;
  background: #fafafa;
  border-left: 1px solid #ddd;
  display: flex;
  flex-direction: column;
  padding: 16px;
  gap: 14px;
  flex-shrink: 0;
  overflow-y: auto;
}
.feedback-panel h2 {
  margin: 0;
  font-size: 1rem;
  color: #333;
  border-bottom: 1px solid #ddd;
  padding-bottom: 8px;
}
.feedback-student-name {
  font-size: 0.85rem;
  color: #555;
  font-style: italic;
  min-height: 1.2em;
}
.feedback-label {
  font-size: 0.85rem;
  font-weight: 600;
  color: #444;
  margin-bottom: 4px;
}
.feedback-panel textarea {
  width: 100%;
  height: 180px;
  resize: vertical;
  border: 1px solid #ccc;
  border-radius: 4px;
  padding: 8px;
  font-size: 0.85rem;
  font-family: Arial, sans-serif;
  line-height: 1.4;
}
.feedback-panel textarea:focus { outline: none; border-color: #5b9bd5; }
.points-row { display: flex; align-items: center; gap: 8px; }
.points-row input[type="text"] {
  width: 80px;
  border: 1px solid #ccc;
  border-radius: 4px;
  padding: 6px 8px;
  font-size: 1rem;
  text-align: center;
}
.points-row input[type="text"]:focus { outline: none; border-color: #5b9bd5; }
.points-row span { font-size: 0.85rem; color: #666; }
.save-btn {
  background: #2563eb;
  color: white;
  border: none;
  border-radius: 4px;
  padding: 8px 16px;
  font-size: 0.9rem;
  cursor: pointer;
  width: 100%;
}
.save-btn:hover { background: #1d4ed8; }
.save-btn:active { background: #1e40af; }
.save-status {
  font-size: 0.8rem;
  min-height: 1.2em;
  text-align: center;
}
.save-status.ok { color: #16a34a; }
.save-status.err { color: #dc2626; }
.export-btn {
  background: #16a34a;
  color: white;
  border: none;
  border-radius: 4px;
  padding: 8px 16px;
  font-size: 0.9rem;
  cursor: pointer;
  width: 100%;
  margin-top: 8px;
}
.export-btn:hover { background: #15803d; }
.export-btn:disabled { background: #9ca3af; cursor: not-allowed; }
.export-status {
  font-size: 0.8rem;
  min-height: 1.2em;
  text-align: center;
  word-break: break-word;
}
.export-status.ok { color: #16a34a; }
.export-status.err { color: #dc2626; }
</style>
</head>
<body>

<!-- Left: student list -->
<div class="sidebar">
  <h2>Students</h2>
  <div id="studentList"></div>
</div>

<!-- Middle: PDF viewer + issue checklist -->
<div class="main">
  <div class="viewer">
    <iframe id="pdfFrame"></iframe>
  </div>
  <div class="checklist" id="checklist">
    <p style="color:#999;font-style:italic;">Select a student to begin.</p>
  </div>
</div>

<!-- Right: feedback panel -->
<div class="feedback-panel">
  <h2>Feedback</h2>
  <div class="feedback-student-name" id="feedbackStudentName">No student selected</div>

  <div>
    <div class="feedback-label">Comments / point deductions</div>
    <textarea id="feedbackComments" placeholder="Describe mistakes and point deductions here..." disabled></textarea>
  </div>

  <div>
    <div class="feedback-label">Points earned</div>
    <div class="points-row">
      <input type="text" id="feedbackPoints" placeholder="—" disabled />
      <span>pts</span>
    </div>
  </div>

  <button class="save-btn" id="saveBtn" onclick="saveFeedback()" disabled>Save</button>
  <div class="save-status" id="saveStatus"></div>

  <hr style="border:none;border-top:1px solid #ddd;margin:4px 0;">
  <button class="export-btn" id="exportBtn" onclick="exportToExcel()">Export All to Excel</button>
  <div class="export-status" id="exportStatus"></div>
</div>

<script>
let currentStudent = null;

function esc(s) {
  if (!s) return '';
  return String(s)
    .replace(/&/g, '&amp;')
    .replace(/</g, '&lt;')
    .replace(/>/g, '&gt;')
    .replace(/"/g, '&quot;');
}

async function loadStudents() {
  const res = await fetch('/api/students');
  const data = await res.json();
  const list = document.getElementById('studentList');
  list.innerHTML = data.students.map(s =>
    '<div class="student" id="student-' + esc(s.name) + '" onclick="selectStudent(this.dataset.name)" data-name="' + esc(s.name) + '">' +
      '<strong>' + esc(s.name) + '</strong><br>' +
      'Written: &#x1F50D; ' + s.written_warnings + ' | \u274c ' + s.written_missing + ' &nbsp; ' +
      'Code: &#x1F50D; ' + s.code_warnings + ' | \u274c ' + s.code_missing +
    '</div>'
  ).join('');
}

function problemSortKey(p) {
  if (p === 'General') return 'zz';
  const m = p.match(/Problem\\s*(\\d+)/i);
  const n = m ? m[1] : '';
  return (n && !isNaN(parseInt(n, 10))) ? ('p' + n.padStart(2, '0')) : p;
}

function renderSection(title, warnings, missing, warningLabel, missingLabel) {
  warningLabel = warningLabel || 'Needs Review';
  missingLabel = missingLabel || 'Not Met';
  const items = [];
  warnings.forEach(r => { items.push({ type: 'warning', problem: r.problem || 'General', r: r }); });
  missing.forEach(r => { items.push({ type: 'missing', problem: r.problem || 'General', r: r }); });
  const byProblem = {};
  items.forEach(it => {
    const p = it.problem;
    if (!byProblem[p]) byProblem[p] = [];
    byProblem[p].push(it);
  });
  const problemOrder = Object.keys(byProblem).sort((a, b) => problemSortKey(a).localeCompare(problemSortKey(b)));
  let html = '<div class="checklist-section"><h3>' + esc(title) + '</h3>';
  if (problemOrder.length === 0) {
    html += '<p class="no-issues">No issues.</p>';
  } else {
    problemOrder.forEach(problem => {
      html += '<div class="problem-group"><h4>' + esc(problem) + '</h4>';
      byProblem[problem].forEach(it => {
        const r = it.r;
        const cls = it.type === 'warning' ? 'warning' : 'missing';
        const label = it.type === 'warning' ? warningLabel : missingLabel;
        html += '<div class="' + cls + '"><span class="problem-badge">' + esc(problem) + '</span><strong>' + esc(r.req_id) + '</strong> <em style="font-size:0.8rem;color:#666;">' + esc(label) + '</em><br>' + esc(r.description) + '</div>';
      });
      html += '</div>';
    });
  }
  html += '</div>';
  return html;
}

async function selectStudent(name) {
  // Highlight selected student
  if (currentStudent) {
    const prev = document.getElementById('student-' + currentStudent);
    if (prev) prev.classList.remove('selected');
  }
  currentStudent = name;
  const el = document.getElementById('student-' + name);
  if (el) el.classList.add('selected');

  // Load student issues
  const res = await fetch('/api/student/' + encodeURIComponent(name));
  const data = await res.json();

  if (data.error) {
    document.getElementById('checklist').innerHTML = '<p class="missing">' + esc(data.error) + '</p>';
    document.getElementById('pdfFrame').src = 'about:blank';
    return;
  }

  if (data.pdf_files.length > 0) {
    document.getElementById('pdfFrame').src = '/pdf/' + encodeURIComponent(data.pdf_files[0].path);
  } else {
    document.getElementById('pdfFrame').src = 'about:blank';
  }

  let html = renderSection('Written / report', data.written.warnings, data.written.missing, 'Needs Review', 'Not Met');
  html += renderSection('Code', data.code.warnings, data.code.missing, 'Needs Review', 'Not Met');
  document.getElementById('checklist').innerHTML = html;

  // Load saved feedback for this student
  const fres = await fetch('/api/feedback/' + encodeURIComponent(name));
  const fdata = await fres.json();

  document.getElementById('feedbackStudentName').textContent = name;
  document.getElementById('feedbackComments').value = fdata.comments || '';
  document.getElementById('feedbackPoints').value = fdata.points || '';
  document.getElementById('feedbackComments').disabled = false;
  document.getElementById('feedbackPoints').disabled = false;
  document.getElementById('saveBtn').disabled = false;
  document.getElementById('saveStatus').textContent = '';
  document.getElementById('saveStatus').className = 'save-status';
}

async function saveFeedback() {
  if (!currentStudent) return;
  const comments = document.getElementById('feedbackComments').value;
  const points = document.getElementById('feedbackPoints').value;
  const statusEl = document.getElementById('saveStatus');

  try {
    const res = await fetch('/api/feedback', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ student: currentStudent, comments, points })
    });
    const data = await res.json();
    if (data.ok) {
      statusEl.textContent = 'Saved \u2713';
      statusEl.className = 'save-status ok';
    } else {
      statusEl.textContent = 'Error: ' + (data.error || 'unknown');
      statusEl.className = 'save-status err';
    }
  } catch (e) {
    statusEl.textContent = 'Network error';
    statusEl.className = 'save-status err';
  }
  setTimeout(() => {
    statusEl.textContent = '';
    statusEl.className = 'save-status';
  }, 3000);
}

async function exportToExcel() {
  const btn = document.getElementById('exportBtn');
  const statusEl = document.getElementById('exportStatus');
  btn.disabled = true;
  statusEl.textContent = 'Exporting...';
  statusEl.className = 'export-status';
  try {
    const res = await fetch('/api/export', { method: 'POST' });
    const data = await res.json();
    statusEl.textContent = data.message || (data.ok ? 'Done' : 'Error');
    statusEl.className = 'export-status ' + (data.ok ? 'ok' : 'err');
  } catch (e) {
    statusEl.textContent = 'Network error';
    statusEl.className = 'export-status err';
  }
  btn.disabled = false;
}

// Allow Ctrl+S / Cmd+S to save while focused in the panel
document.addEventListener('keydown', function(e) {
  if ((e.ctrlKey || e.metaKey) && e.key === 's') {
    if (currentStudent) {
      e.preventDefault();
      saveFeedback();
    }
  }
});

loadStudents();
</script>

</body>
</html>
"""


# ----------------------------
# Main
# ----------------------------

def _find_one(directory, pattern, label):
    """Glob for exactly one match; exit with an error if zero or more than one found."""
    matches = sorted(directory.glob(pattern))
    if len(matches) == 0:
        sys.exit(f"Error: no {label} found in {directory} (looked for '{pattern}')")
    if len(matches) > 1:
        names = ", ".join(m.name for m in matches)
        sys.exit(
            f"Error: multiple {label} files found in {directory} — ambiguous.\n"
            f"  Matches: {names}\n"
            f"  Rename or remove extras so only one matches '{pattern}'."
        )
    return matches[0]


def _find_submissions_dir(directory):
    """Find exactly one subdirectory whose name contains 'submissions'."""
    matches = [p for p in directory.iterdir() if p.is_dir() and "submissions" in p.name.lower()]
    if len(matches) == 0:
        sys.exit(
            f"Error: no submissions directory found in {directory}\n"
            f"  Expected a subdirectory whose name contains 'submissions'."
        )
    if len(matches) > 1:
        names = ", ".join(m.name for m in matches)
        sys.exit(
            f"Error: multiple submissions directories found in {directory} — ambiguous.\n"
            f"  Matches: {names}"
        )
    return matches[0]


def _find_rubric(directory):
    """Find the rubric JSON: prefer files starting with 'rubric', fall back to any *.json."""
    rubric_matches = sorted(directory.glob("rubric*.json"))
    if len(rubric_matches) == 1:
        return rubric_matches[0]
    if len(rubric_matches) > 1:
        names = ", ".join(m.name for m in rubric_matches)
        sys.exit(
            f"Error: multiple rubric JSON files found in {directory} — ambiguous.\n"
            f"  Matches: {names}\n"
            f"  Rename or remove extras so only one matches 'rubric*.json'."
        )
    # No rubric* match — try any *.json
    return _find_one(directory, "*.json", "rubric JSON")


def main():
    parser_cli = argparse.ArgumentParser(
        description="Review grader output alongside student PDFs.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "The directory must contain:\n"
            "  *_written.csv        — written results\n"
            "  *_code.csv           — code results\n"
            "  rubric*.json         — rubric (or *.json if none start with 'rubric')\n"
            "  *submissions* dir    — student submission folders\n"
        ),
    )
    parser_cli.add_argument(
        "directory",
        help="Directory containing the results CSVs, rubric JSON, and submissions folder",
    )
    parser_cli.add_argument("--port", type=int, default=8000, help="Port (default: 8000)")
    parser_cli.add_argument(
        "--feedback-csv",
        default=None,
        help="Path for the feedback CSV (default: <directory>/feedback.csv)",
    )
    parser_cli.add_argument(
        "--grading-xlsx",
        default=None,
        help="Path to PHYS515_grading.xlsx for exporting final grades",
    )
    parser_cli.add_argument(
        "--assignment-num",
        type=int,
        default=None,
        help="Assignment number (e.g. 9) used to find/create the correct columns in the grading xlsx",
    )
    args = parser_cli.parse_args()

    directory = Path(args.directory).resolve()
    if not directory.is_dir():
        sys.exit(f"Error: '{directory}' is not a directory or does not exist.")

    written_csv   = _find_one(directory, "*_written.csv", "written results CSV")
    code_csv      = _find_one(directory, "*_code.csv",    "code results CSV")
    rubric_path   = _find_rubric(directory)
    submissions   = _find_submissions_dir(directory)
    feedback_csv  = Path(args.feedback_csv) if args.feedback_csv else directory / "feedback.csv"

    print(f"  Written CSV  : {written_csv}")
    print(f"  Code CSV     : {code_csv}")
    print(f"  Rubric       : {rubric_path}")
    print(f"  Submissions  : {submissions}")
    print(f"  Feedback CSV : {feedback_csv}")

    parser = ResultsParser(
        written_csv,
        code_csv,
        rubric_path,
        submissions,
        feedback_csv=feedback_csv,
        grading_xlsx=args.grading_xlsx,
        assignment_num=args.assignment_num,
    )
    parser.parse()

    def handler(*args, **kwargs):
        return ReviewerHandler(*args, parser=parser, **kwargs)

    port = args.port
    server = HTTPServer(("localhost", port), handler)

    url = f"http://localhost:{port}/"
    print(f"\nStarting reviewer at {url}")

    threading.Thread(target=lambda: webbrowser.open(url), daemon=True).start()

    server.serve_forever()


if __name__ == "__main__":
    main()
