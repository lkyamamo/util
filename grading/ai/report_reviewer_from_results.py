#!/usr/bin/env python3
"""
Local web UI to review *report* grading output from ``report_grader.py`` alongside
the submitted PDF/DOCX files.

Expects the same directory layout as ``report_grader.list_submission_units``:

  * Top-level ``.pdf`` / ``.docx`` in ``submissions/`` (one file per submission), and/or
  * One subfolder per student with reports inside (scanned recursively).

The project folder should contain:

  * exactly one file matching ``*_report_feedback.json`` (the array written by
    ``report_grader.py``);
  * exactly one subdirectory whose name contains ``submissions`` (or pass
    ``--submissions``);

Usage:
    python report_reviewer_from_results.py /path/to/project1
    python report_reviewer_from_results.py /path/to/project1 --port 8001
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import threading
import urllib.parse
import webbrowser
from http.server import HTTPServer, SimpleHTTPRequestHandler
from pathlib import Path

# Same-directory imports (script may be run as ``python report_reviewer_from_results.py``)
_AIDIR = Path(__file__).resolve().parent
if str(_AIDIR) not in sys.path:
    sys.path.insert(0, str(_AIDIR))

from report_grader import _report_file_kind, list_submission_units
from reviewer_from_results import FeedbackStore

try:
    import openpyxl

    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

# Extracts "Sydney Leggio" from "362403-553690 - Sydney Leggio - Apr 6, 2026 1239 PM"
_NAME_FROM_FOLDER = re.compile(
    r"^\d[\w-]*\s+-\s+(.+?)\s+-\s+\w+\s+\d+,", re.I
)


# ----------------------------
# Report results + file discovery
# ----------------------------


def load_report_feedback_json(path: Path) -> dict[str, dict]:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError(f"Expected a JSON array in {path}")
    out: dict[str, dict] = {}
    for row in data:
        if not isinstance(row, dict):
            continue
        sid = (row.get("student") or "").strip()
        if sid:
            out[sid] = row
    return out


def _find_report_files(submissions_root: Path, path_str: str | None) -> list[dict]:
    """
    Return [{name, path, kind}] where ``path`` is relative to ``submissions_root``,
    suitable for URL paths. ``kind`` is \"pdf\" or \"docx\".
    """
    if not path_str:
        return []
    p = Path(path_str)
    if not p.is_file() and not p.is_dir():
        return []
    root = submissions_root.resolve()
    try:
        p.resolve().relative_to(root)
    except ValueError:
        return []

    out: list[dict] = []
    if p.is_file():
        name = p.name
        k = _report_file_kind(name)
        if not k:
            return []
        rel = str(p.resolve().relative_to(root))
        out.append({"name": name, "path": rel, "kind": k})
        return out

    for f in sorted(p.rglob("*")):
        if not f.is_file():
            continue
        k = _report_file_kind(f.name)
        if not k:
            continue
        rel = str(f.resolve().relative_to(root))
        out.append({"name": f.name, "path": rel, "kind": k})
    out.sort(key=lambda x: (0 if x["kind"] == "pdf" else 1, x["name"].lower()))
    return out


def _safe_file_under_root(root: Path, relpath: str) -> Path | None:
    """Return resolved file path if it exists, is a file, and lies under root."""
    if not relpath or relpath.startswith("..") or relpath.startswith("/"):
        return None
    try:
        full = (root / relpath).resolve()
        full.relative_to(root.resolve())
    except ValueError:
        return None
    if full.is_file():
        return full
    return None


class ReportResultsParser:
    """
    Merges on-disk submission units (``list_submission_units``) with
    ``report_grader`` JSON records. Per-submission id is the file or folder name,
    matching ``report_grader`` output.
    """

    def __init__(
        self,
        report_json: Path,
        submissions_root: Path,
        feedback_csv: Path,
        grading_xlsx: Path | None = None,
        assignment_num: int | None = None,
    ):
        self.report_json = Path(report_json)
        self.submissions_root = Path(submissions_root)
        self.feedback = FeedbackStore(feedback_csv)
        self.grading_xlsx = Path(grading_xlsx) if grading_xlsx else None
        self.assignment_num = assignment_num

        self._by_student: dict[str, dict] = {}
        self._id_to_path: dict[str, str | None] = {}
        self._order: list[str] = []

    def parse(self) -> None:
        self._by_student = load_report_feedback_json(self.report_json)
        units = list_submission_units(str(self.submissions_root))
        self._id_to_path = {u[0]: u[1] for u in units}
        order: list[str] = [u[0] for u in units]
        for sid in sorted(self._by_student.keys()):
            if sid not in self._id_to_path:
                self._id_to_path[sid] = None
                order.append(sid)
        self._order = order

    def record_for(self, student_id: str) -> dict | None:
        return self._by_student.get(student_id)

    def find_report_files(self, student_id: str) -> list[dict]:
        return _find_report_files(
            self.submissions_root, self._id_to_path.get(student_id)
        )

    def export_to_excel(self):
        """Write saved feedback to an xlsx file (same behavior as the code reviewer)."""
        if not _OPENPYXL_AVAILABLE:
            return False, "openpyxl is not installed (pip install openpyxl)"
        if not self.feedback._data:
            return False, "No feedback data to export."
        output_path = self.grading_xlsx
        if not output_path:
            output_path = self.feedback.csv_path.with_suffix(".xlsx")
        if output_path.exists() and self.assignment_num:
            return self._merge_into_existing_xlsx(output_path)
        return self._create_new_xlsx(output_path)

    def _create_new_xlsx(self, output_path: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grades"
        ws.cell(1, 1).value = "Name"
        ws.cell(1, 2).value = "Points"
        ws.cell(1, 3).value = "Comments"
        row = 2
        for folder_name, entry in sorted(self.feedback._data.items()):
            m = _NAME_FROM_FOLDER.match(folder_name)
            student_name = m.group(1).strip() if m else folder_name.strip()
            ws.cell(row, 1).value = student_name
            points = entry.get("points", "").strip()
            if points:
                try:
                    ws.cell(row, 2).value = float(points)
                except ValueError:
                    ws.cell(row, 2).value = points
            ws.cell(row, 3).value = entry.get("comments", "").strip()
            row += 1
        wb.save(str(output_path))
        return True, f"Exported {row - 2} student(s) to {output_path.name}."

    def _merge_into_existing_xlsx(self, output_path: Path):
        point_col_name = f"Sheet {self.assignment_num} Point Grade"
        text_col_name = f"Sheet {self.assignment_num} Text Grade"
        wb = openpyxl.load_workbook(str(output_path))
        ws = wb.active
        header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_map = {str(v).strip(): i + 1 for i, v in enumerate(header_row) if v is not None}
        if point_col_name not in col_map:
            next_col = ws.max_column + 1
            ws.cell(1, next_col).value = point_col_name
            col_map[point_col_name] = next_col
        if text_col_name not in col_map:
            next_col = ws.max_column + 1
            ws.cell(1, next_col).value = text_col_name
            col_map[text_col_name] = next_col
        point_col = col_map[point_col_name]
        text_col = col_map[text_col_name]
        name_col = col_map.get("Name", 1)
        name_to_row: dict[str, int] = {}
        for r in range(2, ws.max_row + 1):
            cell_val = ws.cell(r, name_col).value
            if cell_val:
                name_to_row[str(cell_val).strip().lower()] = r
        written = 0
        unmatched: list[str] = []
        for folder_name, entry in self.feedback._data.items():
            m = _NAME_FROM_FOLDER.match(folder_name)
            student_name = m.group(1).strip() if m else folder_name.strip()
            key = student_name.lower()
            if key not in name_to_row:
                key = next(
                    (k for k in name_to_row if key in k or k in key),
                    None,
                )
            if not key or key not in name_to_row:
                unmatched.append(folder_name)
                continue
            r = name_to_row[key]
            points = entry.get("points", "").strip()
            comments = entry.get("comments", "").strip()
            if points:
                try:
                    ws.cell(r, point_col).value = float(points)
                except ValueError:
                    ws.cell(r, point_col).value = points
            if comments:
                ws.cell(r, text_col).value = comments
            written += 1
        wb.save(str(output_path))
        msg = f"Exported {written} student(s) to {output_path.name}."
        if unmatched:
            msg += f" Could not match: {', '.join(unmatched)}"
        return True, msg


# ----------------------------
# HTTP
# ----------------------------


class ReportReviewerHandler(SimpleHTTPRequestHandler):
    def __init__(self, *args, parser: ReportResultsParser, **kwargs):
        self._parser: ReportResultsParser = parser
        super().__init__(*args, **kwargs)

    @staticmethod
    def _generate_html() -> str:
        return _html_page()

    def do_GET(self) -> None:
        path = urllib.parse.urlparse(self.path).path
        if path in ("/", "/index.html"):
            self._send_html(self._generate_html())
            return
        if path.startswith("/api/students"):
            self._send_json(self._get_students_data())
            return
        if path.startswith("/api/student/"):
            name = urllib.parse.unquote(path.split("/api/student/")[-1])
            self._send_json(self._get_student_data(name))
            return
        if path.startswith("/api/feedback/"):
            name = urllib.parse.unquote(path.split("/api/feedback/")[-1])
            self._send_json(self._parser.feedback.get(name))
            return
        if path.startswith("/pdf/"):
            self._serve_file(path[5:], ("application/pdf", False))
            return
        if path.startswith("/file/"):
            self._serve_file(path[6:], (None, True))
            return
        self.send_response(404)
        self.end_headers()

    def do_POST(self) -> None:
        path = urllib.parse.urlparse(self.path).path
        if path == "/api/feedback":
            length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(length)
            try:
                payload = json.loads(body.decode("utf-8"))
                student = (payload.get("student") or "").strip()
                points = str(payload.get("points") or "").strip()
                comments = str(payload.get("comments") or "").strip()
                if not student:
                    self._send_json({"ok": False, "error": "missing student"})
                    return
                self._parser.feedback.set(student, points, comments)
                self._send_json({"ok": True})
            except Exception as e:
                self._send_json({"ok": False, "error": str(e)})
        elif path == "/api/export":
            try:
                ok, msg = self._parser.export_to_excel()
                self._send_json({"ok": ok, "message": msg})
            except Exception as e:
                self._send_json({"ok": False, "message": str(e)})
        else:
            self.send_response(404)
            self.end_headers()

    def _send_html(self, html: str) -> None:
        self.send_response(200)
        self.send_header("Content-type", "text/html; charset=utf-8")
        self.end_headers()
        self.wfile.write(html.encode("utf-8"))

    def _send_json(self, data: object) -> None:
        self.send_response(200)
        self.send_header("Content-type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode("utf-8"))

    def _serve_file(
        self, path_raw: str, mode: tuple[str | None, bool]
    ) -> None:
        content_type, as_attachment = mode
        rel = urllib.parse.unquote(path_raw)
        full = _safe_file_under_root(self._parser.submissions_root, rel)
        if not full:
            self.send_response(404)
            self.end_headers()
            return
        suffix = full.suffix.lower()
        if suffix == ".pdf" and not as_attachment:
            ct = "application/pdf"
        elif suffix == ".docx":
            ct = (
                "application/vnd.openxmlformats-officedocument."
                "wordprocessingml.document"
            )
        else:
            ct = content_type or "application/octet-stream"
        self.send_response(200)
        self.send_header("Content-type", ct)
        if as_attachment or suffix == ".docx":
            self.send_header(
                "Content-Disposition",
                f'attachment; filename="{full.name}"',
            )
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        with open(full, "rb") as f:
            self.wfile.write(f.read())

    def _get_students_data(self) -> dict:
        p = self._parser
        out: list[dict] = []
        for sid in p._order:
            rec = p.record_for(sid)
            fb = (rec or {}).get("feedback")
            parse_error = bool(isinstance(fb, dict) and "error" in fb)
            nchars = (rec or {}).get("report_chars", 0) if rec else 0
            path = p._id_to_path.get(sid)
            on_disk = bool(path) and Path(path).exists()
            out.append(
                {
                    "name": sid,
                    "in_results": rec is not None,
                    "report_chars": nchars,
                    "parse_error": parse_error,
                    "file_missing": not on_disk,
                }
            )
        return {"students": out}

    @staticmethod
    def _norm_feedback(fb: object) -> dict:
        if not isinstance(fb, dict):
            return {
                "error": "invalid_feedback",
                "summary": "Feedback record is not a valid object.",
            }
        if "error" in fb and fb.get("error") == "parse_failed":
            return dict(fb)
        return {
            "summary": fb.get("summary", ""),
            "strengths": _list_str(fb.get("strengths")),
            "areas_for_improvement": _list_str(fb.get("areas_for_improvement")),
            "guidelines_alignment": str(fb.get("guidelines_alignment", "") or ""),
            "suggested_next_steps": _list_str(fb.get("suggested_next_steps")),
            "error": fb.get("error"),
            "raw_model_output": fb.get("raw_model_output"),
        }

    def _get_student_data(self, student_id: str) -> dict:
        p = self._parser
        if student_id not in p._order:
            return {"error": "Submission not found"}
        files = p.find_report_files(student_id)
        rec = p.record_for(student_id)
        raw_fb = (rec or {}).get("feedback")
        nchars = (rec or {}).get("report_chars", 0) if rec else 0
        if rec is None:
            feedback = {
                "summary": "No entry in the report feedback JSON for this id. Re-run report_grader or check the file name matches.",
                "strengths": [],
                "areas_for_improvement": [],
                "guidelines_alignment": "",
                "suggested_next_steps": [],
            }
        else:
            feedback = self._norm_feedback(raw_fb)
        first_pdf = next((f for f in files if f["kind"] == "pdf"), None)
        first_docx = next((f for f in files if f["kind"] == "docx"), None)
        return {
            "name": student_id,
            "report_files": files,
            "first_pdf": first_pdf,
            "first_docx": first_docx,
            "report_chars": nchars,
            "feedback": feedback,
        }


def _list_str(x: object) -> list[str]:
    if not isinstance(x, list):
        return []
    return [str(i) for i in x if i is not None]


# Embedded HTML/JS
def _html_page() -> str:
    return r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<title>Report grading reviewer</title>
<style>
* { box-sizing: border-box; }
body { font-family: system-ui, Segoe UI, Arial, sans-serif; display: flex; height: 100vh; margin: 0; overflow: hidden; }
.sidebar { width: 300px; min-width: 220px; background: #f4f4f5; overflow-y: auto; padding: 10px; border-right: 1px solid #ddd; flex-shrink: 0; }
.sidebar h2 { margin: 0 0 10px 0; font-size: 1rem; color: #27272a; }
.student { padding: 8px; cursor: pointer; border-bottom: 1px solid #e4e4e7; font-size: 0.9rem; }
.student:hover { background: #e4e4e7; }
.student.selected { background: #bfdbfe; }
.badge { font-size: 0.7rem; color: #71717a; }
.main { flex: 1; display: flex; flex-direction: column; min-width: 0; overflow: hidden; }
.viewer { flex: 1; min-height: 180px; background: #fafafa; }
.viewer iframe { width: 100%; height: 100%; border: none; }
.docx-hint { padding: 12px; background: #fffbeb; border-bottom: 1px solid #fde68a; font-size: 0.9rem; color: #854d0e; }
.docx-hint a { color: #1d4ed8; }
.ai-panel { flex: 1; overflow-y: auto; padding: 14px 16px; border-top: 1px solid #ddd; font-size: 0.9rem; line-height: 1.45; }
.ai-panel h3 { margin: 0 0 8px 0; font-size: 0.85rem; color: #3f3f46; text-transform: uppercase; letter-spacing: 0.03em; }
.ai-block { margin-bottom: 16px; }
.ai-block ul { margin: 6px 0 0 1.1em; padding: 0; }
.raw-out { font-size: 0.75rem; color: #52525b; max-height: 120px; overflow: auto; background: #f4f4f5; padding: 8px; border-radius: 4px; white-space: pre-wrap; word-break: break-word; }
.err-banner { background: #fef2f2; color: #991b1b; padding: 8px; border-radius: 4px; margin-bottom: 12px; font-size: 0.85rem; }
.feedback-panel { width: 300px; min-width: 240px; background: #fafafa; border-left: 1px solid #ddd; display: flex; flex-direction: column; padding: 16px; gap: 12px; flex-shrink: 0; overflow-y: auto; }
.feedback-panel h2 { margin: 0; font-size: 1rem; color: #27272a; border-bottom: 1px solid #e4e4e7; padding-bottom: 8px; }
.feedback-label { font-size: 0.85rem; font-weight: 600; color: #3f3f46; }
.feedback-panel textarea { width: 100%; height: 160px; resize: vertical; border: 1px solid #d4d4d8; border-radius: 4px; padding: 8px; font-size: 0.85rem; }
.points-row { display: flex; align-items: center; gap: 8px; }
.points-row input { width: 80px; border: 1px solid #d4d4d8; border-radius: 4px; padding: 6px; text-align: center; }
.save-btn, .export-btn { color: #fff; border: none; border-radius: 4px; padding: 8px; font-size: 0.9rem; cursor: pointer; width: 100%; }
.save-btn { background: #2563eb; }
.save-btn:hover { background: #1d4ed8; }
.export-btn { background: #16a34a; }
.export-btn:hover { background: #15803d; }
.save-status, .export-status { font-size: 0.8rem; min-height: 1.2em; text-align: center; }
.save-status.ok, .export-status.ok { color: #16a34a; }
.save-status.err, .export-status.err { color: #dc2626; }
</style>
</head>
<body>
<div class="sidebar">
  <h2>Submissions</h2>
  <div id="studentList"></div>
</div>
<div class="main">
  <div class="docx-hint" id="docxHint" style="display:none;"></div>
  <div class="viewer"><iframe id="pdfFrame" title="Report PDF"></iframe></div>
  <div class="ai-panel" id="aiPanel"><p class="badge">Select a submission.</p></div>
</div>
<div class="feedback-panel">
  <h2>Your notes</h2>
  <div class="feedback-label" id="fbName" style="font-size:0.8rem;color:#71717a;">—</div>
  <div><div class="feedback-label">Points</div><div class="points-row"><input type="text" id="pts" disabled/><span>pts</span></div></div>
  <div><div class="feedback-label">Comments / deductions</div><textarea id="com" disabled placeholder="Instructor comments…"></textarea></div>
  <button type="button" class="save-btn" id="saveBtn" onclick="saveFb()" disabled>Save</button>
  <div class="save-status" id="saveSt"></div>
  <hr style="border:none;border-top:1px solid #e4e4e7;"/>
  <button class="export-btn" id="exBtn" onclick="exportX()">Export to Excel</button>
  <div class="export-status" id="exSt"></div>
</div>
<script>
const esc = s => !s ? '' : String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
let current = null;

async function loadList() {
  const res = await fetch('/api/students');
  const data = await res.json();
  const el = document.getElementById('studentList');
  el.innerHTML = (data.students || []).map(s => {
    let tags = [];
    if (!s.in_results) tags.push('no JSON');
    if (s.parse_error) tags.push('parse err');
    if (s.file_missing) tags.push('no file on disk');
    const t = tags.length ? ' <span class="badge">(' + esc(tags.join(', ')) + ')</span>' : '';
    return '<div class="student" data-name="' + esc(s.name) + '"><strong>' + esc(s.name) + '</strong><br><span class="badge">' + (s.report_chars || 0) + ' chars</span>' + t + '</div>';
  }).join('');
  el.querySelectorAll('.student').forEach(d => d.addEventListener('click', () => select(d.getAttribute('data-name'))));
}

function renderAI(f) {
  if (f.error === 'parse_failed' && f.raw_model_output) {
    return '<div class="err-banner">Model output was not valid JSON. Snippet below.</div><div class="raw-out">' + esc(f.raw_model_output) + '</div>';
  }
  if (f.error) {
    return '<div class="err-banner">Error: ' + esc(f.error) + '</div>';
  }
  let h = '<div class="ai-block"><h3>Summary</h3><p>' + esc(f.summary || '—') + '</p></div>';
  h += '<div class="ai-block"><h3>Strengths</h3><ul>' + (f.strengths || []).map(x => '<li>' + esc(x) + '</li>').join('') + '</ul></div>';
  h += '<div class="ai-block"><h3>Areas for improvement</h3><ul>' + (f.areas_for_improvement || []).map(x => '<li>' + esc(x) + '</li>').join('') + '</ul></div>';
  h += '<div class="ai-block"><h3>Guidelines</h3><p>' + esc(f.guidelines_alignment || '—') + '</p></div>';
  h += '<div class="ai-block"><h3>Next steps</h3><ul>' + (f.suggested_next_steps || []).map(x => '<li>' + esc(x) + '</li>').join('') + '</ul></div>';
  return h;
}

function highlightRow(name) {
  document.querySelectorAll('.student').forEach(function(el) {
    el.classList.toggle('selected', el.getAttribute('data-name') === name);
  });
}
async function select(name) {
  current = name;
  highlightRow(name);
  const res = await fetch('/api/student/' + encodeURIComponent(name));
  const d = await res.json();
  if (d.error) {
    document.getElementById('aiPanel').innerHTML = '<p class="err-banner">' + esc(d.error) + '</p>';
    return;
  }
  document.getElementById('aiPanel').innerHTML = renderAI(d.feedback);
  const hint = document.getElementById('docxHint');
  const frame = document.getElementById('pdfFrame');
  if (d.first_pdf) {
    frame.src = '/pdf/' + encodeURIComponent(d.first_pdf.path);
    hint.style.display = 'none';
  } else if (d.first_docx) {
    frame.src = 'about:blank';
    hint.style.display = 'block';
    hint.innerHTML = 'No PDF preview. <a href="/file/' + encodeURIComponent(d.first_docx.path) + '">Download Word file</a> (' + esc(d.first_docx.name) + ').';
  } else {
    frame.src = 'about:blank';
    hint.style.display = 'block';
    hint.textContent = 'No PDF or Word file found for this id under the submissions folder.';
  }
  const f = await (await fetch('/api/feedback/' + encodeURIComponent(name))).json();
  document.getElementById('fbName').textContent = name;
  document.getElementById('pts').value = f.points || '';
  document.getElementById('com').value = f.comments || '';
  ['pts','com','saveBtn'].forEach(id => { document.getElementById(id).disabled = false; });
  document.getElementById('saveSt').textContent = '';
}

async function saveFb() {
  if (!current) return;
  const st = document.getElementById('saveSt');
  try {
    const r = await fetch('/api/feedback', { method: 'POST', headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ student: current, points: document.getElementById('pts').value, comments: document.getElementById('com').value }) });
    const j = await r.json();
    st.textContent = j.ok ? 'Saved ✓' : (j.error || 'Error');
    st.className = 'save-status ' + (j.ok ? 'ok' : 'err');
  } catch (e) { st.textContent = 'Network error'; st.className = 'save-status err'; }
  setTimeout(() => { st.textContent = ''; st.className = 'save-status'; }, 2500);
}

async function exportX() {
  const b = document.getElementById('exBtn');
  const st = document.getElementById('exSt');
  b.disabled = true; st.textContent = '…';
  try {
    const r = await fetch('/api/export', { method: 'POST' });
    const j = await r.json();
    st.textContent = j.message || '';
    st.className = 'export-status ' + (j.ok ? 'ok' : 'err');
  } catch (e) { st.textContent = e.message; st.className = 'export-status err'; }
  b.disabled = false;
}

document.addEventListener('keydown', e => { if ((e.metaKey || e.ctrlKey) && e.key === 's' && current) { e.preventDefault(); saveFb(); } });
loadList();
</script>
</body>
</html>"""


# ----------------------------
# CLI
# ----------------------------


def _find_one(directory: Path, pattern: str, label: str) -> Path:
    matches = sorted(directory.glob(pattern))
    if len(matches) == 0:
        sys.exit(
            f"Error: no {label} found in {directory} (looked for '{pattern}')"
        )
    if len(matches) > 1:
        names = ", ".join(m.name for m in matches)
        sys.exit(
            f"Error: multiple {label} in {directory}.\n  Matches: {names}"
        )
    return matches[0]


def _find_submissions_dir(directory: Path) -> Path:
    matches = [
        p
        for p in directory.iterdir()
        if p.is_dir() and "submissions" in p.name.lower()
    ]
    if len(matches) == 0:
        sys.exit(
            f"Error: no submissions directory under {directory}\n"
            "  Expected a subdirectory whose name contains 'submissions'."
        )
    if len(matches) > 1:
        names = ", ".join(m.name for m in matches)
        sys.exit(
            f"Error: multiple submissions directories: {names}"
        )
    return matches[0]


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Review report_grader output next to student PDFs/DOCX.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument(
        "directory",
        help="Folder containing *_report_feedback.json and a *submissions* subfolder",
    )
    ap.add_argument("--port", type=int, default=8000, help="HTTP port (default 8000)")
    ap.add_argument(
        "--submissions",
        default=None,
        help="Path to submissions root (default: the sole *submissions* subdir)",
    )
    ap.add_argument(
        "--results-json",
        default=None,
        help="Path to *_report_feedback.json (default: only match in directory)",
    )
    ap.add_argument(
        "--feedback-csv",
        default=None,
        help="Path for instructor notes CSV (default: <dir>/report_reviewer_feedback.csv)",
    )
    ap.add_argument(
        "--grading-xlsx",
        default=None,
        help="Path to a master grading workbook for export merge",
    )
    ap.add_argument(
        "--assignment-num",
        type=int,
        default=None,
        help="Sheet number for merge columns in the grading xlsx",
    )
    args = ap.parse_args()

    base = Path(args.directory).resolve()
    if not base.is_dir():
        sys.exit(f"Error: not a directory: {base}")

    jpath = (
        Path(args.results_json).resolve()
        if args.results_json
        else _find_one(base, "*_report_feedback.json", "report feedback JSON")
    )
    sub_root = (
        Path(args.submissions).resolve()
        if args.submissions
        else _find_submissions_dir(base)
    )
    feedback_csv = (
        Path(args.feedback_csv).resolve()
        if args.feedback_csv
        else base / "report_reviewer_feedback.csv"
    )
    gxl = Path(args.grading_xlsx).resolve() if args.grading_xlsx else None

    print(f"  Report JSON   : {jpath.name}")
    print(f"  Submissions   : {sub_root}")
    print(f"  Feedback CSV  : {feedback_csv}")

    p = ReportResultsParser(
        jpath,
        sub_root,
        feedback_csv=feedback_csv,
        grading_xlsx=gxl,
        assignment_num=args.assignment_num,
    )
    p.parse()

    def handler_factory(
        *a, **kw
    ) -> ReportReviewerHandler:
        return ReportReviewerHandler(*a, parser=p, **kw)

    port = args.port
    httpd = HTTPServer(("127.0.0.1", port), handler_factory)
    url = f"http://127.0.0.1:{port}/"
    print(f"\nReport reviewer: {url}")
    threading.Thread(target=lambda: webbrowser.open(url), daemon=True).start()
    httpd.serve_forever()


if __name__ == "__main__":
    main()
